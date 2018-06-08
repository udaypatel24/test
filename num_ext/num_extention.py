#Python script to extend the number
import openpyxl
from openpyxl import Workbook
from time import  strftime
import time
import random
import getpass

print ("---------------------------------------------------------")
print ("---------------------------------------------------------")
print ("---------------------------------------------------------")
print ("---------------------------------------------------------")
print ("--------------------   UDAY     -------------------------")
print ("---------------------------------------------------------")
print ("---------------------------------------------------------")
print ("---------------------------------------------------------")
print ("---------------------------------------------------------")

password = getpass.getpass("Please enter your password :  ")
if ( password == "iojknm"):
    wb = openpyxl.load_workbook('ext.xlsx')
    wb1 = Workbook()
    ws1 = wb1.active
    sheet = wb['number']
    name = strftime("%Y-%m-%d %H_%M_%S")
    for n in range(1, 20):
        rown = sheet[n]
        x = 1
        y = n
        for i in rown:
            myval = i.value
            myvals = str(myval)
            if "-" in myvals:
                n1,n2 = myvals.split("-")
                num1 = int(n1)
                num2 = int(n2)
                num3 = num2 + 1
                for i in range (num1, num3):
                    ws1.cell(row=x, column=y, value=i)
                    x+=1
            
            else:
#                val = int(myvals)
                ws1.cell(row=x, column=y, value=myval)
                x+=1
    wb1.save("results_" + name + ".xlsx")
    
else:
    print("You entered wrong password . Good Bye !!!!!!!")
    time.sleep(5)
   
